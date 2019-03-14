import json
import random
import openpyxl








def write07Excel(path):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'HundredPoems'

    value = [["title", "url", "content"]
             ]
    poems = json.load(open("seged_poem.json", "r", encoding="utf-8"))
    random.shuffle(poems)
    poems = poems[:100]
    for ele in poems:
        title = ele["poem_title"]
        url = ele["url"]
        content = ele["origin_poem"]
        value.append([title,url,content])

    for i in range(0, len(value)):
        for j in range(0, len(value[i])):
            sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))

    wb.save(path)
    print("写入数据成功！")

def read07Excel(path,sheet_name):
    wb = openpyxl.load_workbook(path)
    sheet = wb.get_sheet_by_name(sheet_name)

    titl_url = {}
    List = []
    for i , row in enumerate( sheet.rows):
        poem_title = row[0].value
        List.append(poem_title)
        url = row[1].value
        if poem_title not in titl_url:
            titl_url[poem_title] = []
        titl_url[poem_title].append(url)
    return titl_url , List

def getData(title_url):
    with open("HundredPoems.txt", "w", encoding="utf-8") as fout:
        for i, ele in enumerate(poems):
            # title = ele['poem_title']
            # url = ele['url']
            # fout.write(title+"\n"+url+"\n")
            # fout.write("\n")
            # for key in ele:
            #     print("key = ",key, " " ,ele[key])
            # paras = ele['paras']

            # for key in paras:
            #     print("key = ", key)
            break
if __name__ == "__main__":
    # write07Excel("HundredPoems.xlsx")
    import os
    print(os.getcwd())
    title_url,List=read07Excel("HundredPoems.xlsx",'HundredPoems')
    # poems = json.load(open("seged_poem.json", "r", encoding="utf-8"))
    poems = json.load(open("processed_poem_2019.json.json", "r", encoding="utf-8"))
    print("len_poems = ",len(poems))
    res = []
    for poem in poems:
        if poem["poem_title"] in title_url and poem["url"] in title_url[poem["poem_title"]]:
            # if poem["poem_title"] == "幸福的理由":
            #     print(poem)
            #     paras = poem['paras']
            #     for p in paras:
            #         print("p = ",p)
            #     exit(89)
            res.append((poem , List.index(poem["poem_title"])))
    print(len(res))
    res.sort(key = lambda t:t[1])
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'HundredParas'
    value = [["title", "url", "para_id", "content"]
             ]
    for i in range(0, len(value)):
        for j in range(0, len(value[i])):
            sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))
    offset = len(value) + 1
    for index,(poem,shuncu ) in enumerate( res ):
        url = poem["url"]
        title = poem["poem_title"]
        paras = poem["paras"]
        if len(paras) == 0:
            print(poem)
            # sheet.cell(row=index + offset + 1, column=1, value=str(title))
            # sheet.cell(row=index + offset + 1, column=2, value=str(url))
            # sheet.cell(row=index + offset + 1, column=3, value=str(0))
            # sheet.cell(row=index + offset + 1, column=4, value=str(poem['origin_poem']))
            sheet.cell(row=offset , column=1, value=str(title))
            sheet.cell(row=offset, column=2, value=str(url))
            sheet.cell(row=offset , column=3, value=str(0))
            sheet.cell(row=offset, column=4, value=str(poem['origin_poem']))
            offset += 1
            continue
        for para_id , para in enumerate( paras):
            para_title = para["para_title"]
            if para_title == "":
                para_title = title
            para_content = para["para_content"]
            # sheet.cell(row=index + offset + 1 + para_id + 1, column= 1, value=str(para_title))
            # sheet.cell(row=index + offset + 1+ para_id + 1, column= 2, value=str(url))
            # sheet.cell(row=index + offset + 1+ para_id + 1, column= 3, value=str(para_id))
            # sheet.cell(row=index + offset + 1+ para_id + 1, column= 4, value=str('\n'.join(para_content)))
            sheet.cell(row=offset , column=1, value=str(para_title))
            sheet.cell(row=offset , column=2, value=str(url))
            sheet.cell(row=offset , column=3, value=str(para_id))
            sheet.cell(row= offset, column=4, value=str('\n'.join(para_content)))
            offset += 1
    wb.save("HundredParas2019.xlsx")





