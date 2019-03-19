'''
1.利用种子词（：爱情）得到相关篇章。
2.从相关篇章中利用TextRank算法抽取关键词，取top_n加入到种子词集合中。
重复1.2.两个步骤，直到达到终止条件
终止条件：得到的新一批的种子词与主题（爱情）并不相关了。or 得到的新的篇章更少了。or 得到的新的种子词更少了。

输入：种子词文件路径
输出：每轮迭代输出，新的（划重点）种子词，以及新的（划重点）相关文档。
'''
import os
import json
import openpyxl
import random
import pickle
import jieba
from jieba import analyse
import re
from Model.Poem.config import opt
from tqdm import tqdm

def write07Excel(path):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = '1'

    value = [["title", "url", "content"]
             ]
    poems = json.load(open("seged_poem.json", "r", encoding="utf-8"))
    random.shuffle(poems)
    poems = poems[:100]
    for ele in poems:
        title = ele["poem_title"]
        url = ele["url"]
        content = ele["origin_poem"]
        value.append([title,url,content])

    for i in range(0, len(value)):
        for j in range(0, len(value[i])):
            sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))

    wb.save(path)
    print("写入数据成功！")

def read07Excel(path,sheet_name):
    wb = openpyxl.load_workbook(path)
    sheet = wb.get_sheet_by_name(sheet_name)

    titl_url = {}
    List = []
    for i , row in enumerate( sheet.rows):
        poem_title = row[0].value
        List.append(poem_title)
        url = row[1].value
        if poem_title not in titl_url:
            titl_url[poem_title] = []
        titl_url[poem_title].append(url)
    return titl_url , List





class GetSeedWord:
    '''
    迭代式的得到种子词
    '''
    def __init__(self , seed_word_path , prose_path , proseSelcted = None , top_n = 10 , max_iter_num = 3 ,threshold = 0.001,ExtractMode="tfidf",idf_path=""):
        '''
        :param seed_word_path: 种子词的文件路径
        :param sheet_name: 种子词excel文件中的sheetName
        :param prose_path: 文章的文件路径
        :param top_n: 取前top_n个种子词
        :param max_iter_num: 最大的迭代次数
        '''
        if seed_word_path is None or not os.path.exists(seed_word_path):
            raise Exception('初始种子词文件必须提供')

        self.seed_word_path = seed_word_path
        self.prose_path = prose_path
        self.proseSelcted = proseSelcted
        self.top_n = top_n
        self.max_iter_num = max_iter_num
        self.threshold = threshold
        self.ExtractMode = ExtractMode
        if ExtractMode == "tfidf" and idf_path:
            analyse.set_idf_path(idf_path)
        self.seed_words = set()
        self.seed_words |= self.loadSeedWord(seed_word_path)

        self.prosePastSelected = set()
        if proseSelcted is not None:
            self.prosePastSelected |= self.loadProseNumber(path=proseSelcted)

        #self.poems = self.getProse(prose_path)
        self.poems = pickle.load(open("poemsIdf.pkl","rb"))
        print("散文加载成功")

        self.proseCurrentSelected = set() #用当前这些关键词检索出来的文章
        self.wordCurrentSelected = set() #用当前的这些文章检索出来的关键词
        self.wordCurrentSelectedDict = dict()

        self.ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
        self.PoemID2Index()
        print("初始化结束")


    def loadSeedWord(self,path):
        '''
        :param path:  Excel的文件路径
        :return: 种子词集合 set type
        '''
        seedWord = set()
        wb = openpyxl.load_workbook(path)
        sheet_names = wb.get_sheet_names()
        for sheet_name in sheet_names:
            sheet = wb.get_sheet_by_name(sheet_name)
            for i, row in enumerate(sheet.rows):
                word = row[0].value
                seedWord.add(word)
        print("种子词读取成功！")
        return seedWord

    def dumpSeedWord(self, path, seedWordTuple):
        '''
        :param path:  Excel的文件路径
        :param sheet_name: Excel的文件路径
        :return: None
        '''
        if os.path.exists(path):
            wb = openpyxl.load_workbook(path)
            sheet_names = wb.get_sheet_names()
            sheet_name = str( len(sheet_names) + 1)
            wb.create_sheet(sheet_name)
            sheet = wb.get_sheet_by_name(sheet_name)
        else:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = '1'
        values = ['单词', '最大权重', '总权重', '文章数', '文章列表']
        for j in range(len(values)):
            sheet.cell(row=1, column=j + 1, value=values[j])
        for i, weightTuple in enumerate(seedWordTuple):
            word = weightTuple[0]
            weightDict = weightTuple[1]
            weight = weightDict['total_weight']
            max_wieght = weightDict['max_weight']
            sheet.cell(row=i + 2, column=1, value=word)
            sheet.cell(row=i + 2, column=2, value=max_wieght)
            sheet.cell(row=i + 2, column=3, value=weight)
            contributors = weightDict['contributors']
            sheet.cell(row=i + 2, column=4, value=len(contributors))
            contributors = sorted(contributors, key=lambda t: t['weight'], reverse=True)
            for index, con_dict in enumerate(contributors):
                sheet.cell(row=i + 2, column=4 + index * 2 + 1, value=con_dict['weight'])
                poem_id = con_dict['poem_id']
                para_id = con_dict['para_id']
                content = '\n'.join(self.poems[self.reverseDict[poem_id]]['paras'][para_id]['para_content'])
                content = self.ILLEGAL_CHARACTERS_RE.sub(r'', content)
                sheet.cell(row=i + 2, column=4 + index * 2 + 2, value=content)
        wb.save(path)
        print("种子词写入成功！")

    def loadProseNumber(self , path):
        '''
        :param path: 存储已经被选择的文章id的excel文件路径
        :return:
        '''
        proseNumbers = set()
        if not os.path.exists(path):
            return proseNumbers
        wb = openpyxl.load_workbook(path)
        sheet_names = wb.get_sheet_names()
        for sheet_name in sheet_names:
            sheet = wb.get_sheet_by_name(sheet_name)
            for i, row in enumerate(sheet.rows):
                if row[0].value is None:
                    break
                poem_id , para_id = int(row[0].value) , int(row[1].value)
                proseNumbers.add((poem_id , para_id))
        print("文章编号读取成功！")
        return proseNumbers

    def PoemID2Index(self):
        self.reverseDict = dict()
        for index , poem in enumerate(self.poems):
            self.reverseDict[poem['poem_id']] = index
        return self.reverseDict

    def dumpProseNumber(self , path , proseNumbers):
        '''
        :param path:  Excel的文件路径
        :return: None
        '''
        if os.path.exists(path):
            wb = openpyxl.load_workbook(path)
            sheet_names = wb.get_sheet_names()
            sheet_name = str(len(sheet_names) + 1)
            wb.create_sheet(sheet_name)
            sheet = wb.get_sheet_by_name(sheet_name)
        else:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = '1'
        for i, (poem_id, para_id) in enumerate(proseNumbers):
            sheet.cell(row=i + 1, column=1, value=poem_id)
            sheet.cell(row=i + 1, column=2, value=para_id)
            content = '\n'.join(self.poems[self.reverseDict[poem_id]]['paras'][para_id]['para_content'])
            content = self.ILLEGAL_CHARACTERS_RE.sub(r'', content)
            sheet.cell(row=i + 1, column=3, value=content)
        wb.save(path)
        print("文章编号写入成功！")

    def getProse(self,path):
        with open(path, "r", encoding="utf8") as fout:
            return json.load(fout)

    def TextRank(self, text ,topK=10, withWeight=True, allowPOS=('ns', 'n', 'vn', 'v')):
        '''
        :param text: string type
        :return: List of Tuples :[(word , float_weight) ,……,() ]
        '''
        return analyse.textrank(text, topK=topK, withWeight=withWeight, allowPOS=allowPOS)

    def TF_IDF(self , text ,topK=10, withWeight=True, allowPOS=('ns', 'n', 'vn', 'v')):
        return analyse.extract_tags(text, topK=topK, withWeight=withWeight, allowPOS=allowPOS)

    def ExtractKeyword(self, text ,topK=10, withWeight=True, allowPOS=('ns', 'n', 'vn', 'v'), mode = "tfidf"):
        if mode == "tfidf":
            return self.TF_IDF(text ,topK=topK, withWeight=withWeight, allowPOS=allowPOS)
        elif mode == "textrank":
            return self.TextRank(text ,topK=topK, withWeight=withWeight, allowPOS=allowPOS)
        else:
            return self.CustomTextRank(text ,topK=topK, withWeight=withWeight, allowPOS=allowPOS)

    def Select(self):
        for poem in tqdm(self.poems):
            for para_id , para in enumerate(poem['paras']):
                key_words_dict  = {}
                if 'temp_key_words' in para:
                    key_words_dict = para['temp_key_words']
                else:
                    all_content = '\n'.join(para['para_content'])
                    word_weight = self.ExtractKeyword(all_content,topK=self.top_n,mode=self.ExtractMode)
                    for word,weight in word_weight:
                        if word not in key_words_dict:
                            key_words_dict[word] = weight
                        else:
                            key_words_dict[word] = max(key_words_dict[word],weight)
                    para['temp_key_words'] = key_words_dict
                this_prose_weight = 0
                for word, weight in key_words_dict.items():
                    if word in self.seed_words:
                        this_prose_weight += weight
                if this_prose_weight >= self.threshold:
                    self.proseCurrentSelected.add((poem['poem_id'],para_id))
                    for word, weight in key_words_dict.items():
                        if word not in self.seed_words:
                            if word not in self.wordCurrentSelectedDict:
                                self.wordCurrentSelectedDict[word] = {"total_weight":0,'max_weight':0,"contributors":[]}
                            self.wordCurrentSelectedDict[word]['total_weight'] += weight
                            self.wordCurrentSelectedDict[word]['max_weight'] = max(weight,self.wordCurrentSelectedDict[word]['max_weight'])
                            self.wordCurrentSelectedDict[word]['contributors'].append({"weight":weight,"poem_id":poem['poem_id'],"para_id":para_id})

        List_Tuple = sorted(self.wordCurrentSelectedDict.items(),key= lambda t:(t[1]['max_weight'],t[1]['total_weight'],-len(t[1]['contributors'])),reverse=True)
        print("Len(List_Tuple) = ",len(List_Tuple))
        for word , weightTuple in List_Tuple:
            self.wordCurrentSelected.add(word)
        return List_Tuple


    def forward(self , iter_num = 3):
        iter_num = min(iter_num, self.max_iter_num)
        for i in range(iter_num):
            print("iter in %d epoch"%i)
            List_Tuple = self.Select() #添加内容到 self.wordCurrentSelected 和 self.proseCurrentSelected 中
            if len(List_Tuple) ==0:
                break
            print("选择完成")

            self.proseCurrentSelected -= self.prosePastSelected
            print("本轮迭代选择文章数:%d"%len(self.proseCurrentSelected))
            self.dumpProseNumber(self.proseSelcted,self.proseCurrentSelected)
            self.prosePastSelected |= self.proseCurrentSelected

            self.wordCurrentSelected -= self.seed_words
            self.dumpSeedWord(self.seed_word_path, List_Tuple)
            self.seed_words |= self.wordCurrentSelected


            self.wordCurrentSelected = set()
            self.proseCurrentSelected = set()
            self.wordCurrentSelectedDict = dict()
        print("迭代结束")
        # for poem in self.poems:
        #     for para in poem["paras"]:
        #         print(para['temp_key_words'])
        pickle.dump(self.poems,open("poemsIdf.pkl","wb"))



if __name__ == "__main__":
    projectPath = "FirstDayOnMS2"
    path = os.path.abspath(os.getcwd())
    path = path.split(projectPath)
    path = os.path.join(path[0],projectPath)
    print("path  = ",path)
    seedGetter = GetSeedWord(seed_word_path=os.path.join(path,"Model\\Poem\\seed_love.xlsx"),
                             prose_path=os.path.join(path,"Data\\Poem\\processed_poem_2019.json"),
                             proseSelcted=os.path.join(path,"Model\\Poem\\seed_love_story.xlsx"),
                             ExtractMode="tfidf",
                             idf_path="../idf.txt",
                             max_iter_num=1)
    seedGetter.forward()


