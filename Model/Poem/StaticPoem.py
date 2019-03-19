import os
import sys
import json
import jieba
import openpyxl
from jieba import analyse

class StaticPoem:
    def __init__(self,prose_path,
                 idf_path = None,
                 top_n = 10,
                 save_path = None):
        super(StaticPoem,self).__init__()
        self.idf_path = idf_path
        self.top_n = top_n
        if idf_path!= None:
            analyse.set_idf_path(idf_path)
        self.poems = json.load(open(prose_path,"r",encoding="utf-8"))
        self.save_path=save_path

    def TextRank(self, text ,topK=10, withWeight=True, allowPOS=('ns', 'n', 'vn', 'v')):
        '''
        :param text: string type
        :return: List of Tuples :[(word , float_weight) ,……,() ]
        '''
        return analyse.textrank(text, topK=topK, withWeight=withWeight, allowPOS=allowPOS)

    def TF_IDF(self , text ,topK=10, withWeight=True, allowPOS=('ns', 'n', 'vn', 'v')):
        return analyse.extract_tags(text, topK=topK, withWeight=withWeight, allowPOS=allowPOS)

    def ExtractKeyword(self, text ,topK=10, withWeight=True, allowPOS=('ns', 'n', 'vn', 'v'), mode = "tfidf"):
        if mode == "tfidf":
            return self.TF_IDF(text ,topK=topK, withWeight=withWeight, allowPOS=allowPOS)
        elif mode == "textrank":
            return self.TextRank(text ,topK=topK, withWeight=withWeight, allowPOS=allowPOS)

    def static(self):
        print("urls = ",len(self.poems))
        prose_num = 0
        webSiteNames = []
        poemClass = {}
        for ele in self.poems:
            prose_num += len(ele['paras'])
            if ele['website_name'] not in webSiteNames:
                webSiteNames.append(ele['website_name'])
            if ele['poem_class'] not in poemClass:
                poemClass[ele['poem_class']] = {'num':len(ele['paras']),'idf_key_words':set(),"textrank_key_words":set(),
                                                "idf_word_weight":[],"textrank_word_weight":[]}
            else:
                poemClass[ele['poem_class']]["num"] += len(ele['paras'])

            for para_id , para in enumerate(ele['paras']):
                all_content = '\n'.join(para['para_content'])
                if "idf_key_words" in para:
                    idf_word_weight = para["idf_key_words"]
                else:
                    idf_word_weight = self.ExtractKeyword(all_content,topK=self.top_n,mode="tfidf")
                if "key_words" in para:
                    text_word_weight = para["key_words"]
                else:
                    text_word_weight = self.ExtractKeyword(all_content,topK=self.top_n,mode="textrank")
                for word_weight in idf_word_weight:
                    if word_weight[0] not in poemClass[ele['poem_class']]["idf_key_words"]:
                        poemClass[ele['poem_class']]["idf_key_words"].add(word_weight[0])
                        poemClass[ele['poem_class']]["idf_word_weight"].append([word_weight[0],word_weight[1]])
                    else:
                        for e_word_weight in poemClass[ele['poem_class']]["idf_word_weight"]:
                            if e_word_weight[0] == word_weight[0]:
                                e_word_weight[1] = max(e_word_weight[1],word_weight[1])
                                break

                for word_weight in text_word_weight:
                    if word_weight[0] not in poemClass[ele['poem_class']]["textrank_key_words"]:
                        poemClass[ele['poem_class']]["textrank_key_words"].add(word_weight[0])
                        poemClass[ele['poem_class']]["textrank_word_weight"].append([word_weight[0],word_weight[1]])
                    else:
                        for e_word_weight in poemClass[ele['poem_class']]["textrank_word_weight"]:
                            if e_word_weight[0] == word_weight[0]:
                                e_word_weight[1] = max(e_word_weight[1],word_weight[1])
                                break
        for ele_key,ele_value_dict in poemClass.items():
            ele_value_dict["idf_word_weight"].sort(key = lambda t:t[1] , reverse=True)
            ele_value_dict["textrank_word_weight"].sort(key = lambda t:t[1] , reverse=True)

        print("poem num = ",prose_num)
        print(webSiteNames)
        print(poemClass)
        self.poemClass = poemClass

    def Write2xlsx(self):
        print("op")
        if self.save_path is None:
            return
        wb = openpyxl.Workbook()
        for className,classDict in self.poemClass.items():
            print("className = ",className)
            sheet = wb.create_sheet(className)
            sheet.cell(row = 1, column=1, value="单词")
            sheet.cell(row = 1, column=2, value="idf")
            sheet.cell(row=1, column=3, value="")
            sheet.cell(row=1, column=4, value="单词")
            sheet.cell(row = 1, column=5, value="textrank")
            offset = 2
            for index ,idf_word_weight in enumerate( classDict["idf_word_weight"] ):
                sheet.cell(row=offset + index, column=1, value=idf_word_weight[0])
                sheet.cell(row=offset + index, column=2, value=idf_word_weight[1])

            for index ,idf_word_weight in enumerate( classDict["textrank_word_weight"] ):
                sheet.cell(row=offset + index, column=4, value=idf_word_weight[0])
                sheet.cell(row=offset + index, column=5, value=idf_word_weight[1])


        wb.save(self.save_path)
        print("写入数据成功！")


if __name__ == "__main__":
    projectPath = "FirstDayOnMS2"
    path = os.path.abspath(os.getcwd())
    path = path.split(projectPath)
    path = os.path.join(path[0],projectPath)
    print("path  = ",path)
    prose_path = os.path.join(path, "Data\\Poem\\processed_poem_2019.json")
    S = StaticPoem(prose_path = os.path.join(path, "Data\\Poem\\processed_poem_2019.json"),
                   save_path="poems.xlsx")
    S.static()
    S.Write2xlsx()